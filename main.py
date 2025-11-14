import os
import io
import calendar
import holidays
import locale
from datetime import date, datetime
from flask import Flask, send_file, request
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

app = Flask(__name__)

# Mapping from Italian month names to month numbers
italian_months = {
    'gennaio': 1, 'febbraio': 2, 'marzo': 3, 'aprile': 4, 'maggio': 5, 'giugno': 6,
    'luglio': 7, 'agosto': 8, 'settembre': 9, 'ottobre': 10, 'novembre': 11, 'dicembre': 12
}

@app.route("/")
def index():
    return send_file('src/index.html')

@app.route("/submit", methods=['POST'])
def submit():
    # Set locale to Italian to get day names in Italian
    try:
        locale.setlocale(locale.LC_TIME, 'it_IT.UTF-8')
    except locale.Error:
        locale.setlocale(locale.LC_TIME, '') # Fallback to default locale

    nome = request.form.get("nome")
    cognome = request.form.get("cognome")
    mese_str = request.form.get("mese", '').lower()
    ferie_str = request.form.get("ferie", "")
    malattia_str = request.form.get("malattia", "")
    permessi_str = request.form.get("permessi", "")

    # --- Data Parsing ---
    try:
        giorni_ferie = set(map(int, ferie_str.split(','))) if ferie_str else set()
        giorni_malattia = set(map(int, malattia_str.split(','))) if malattia_str else set()
        
        permessi = {}
        if permessi_str:
            for item in permessi_str.replace(' ', '').split(','):
                if not item:
                    continue
                day_str, time_range = item.split(':', 1)
                day = int(day_str)
                start_time_str, end_time_str = time_range.split('-')
                
                start_time = datetime.strptime(start_time_str, '%H:%M')
                end_time = datetime.strptime(end_time_str, '%H:%M')
                
                duration = (end_time - start_time).total_seconds() / 3600
                permessi[day] = {"range": time_range, "hours": duration}

    except ValueError:
        return "Formato giorni o permessi non valido. Usare i formati corretti (es. Ferie: 1,5. Permessi: 13:10:00-11:30)", 400

    current_year = date.today().year
    month_number = italian_months.get(mese_str)
    if not month_number:
        return "Mese non valido. Per favore inserisci un mese in italiano.", 400

    _, num_days = calendar.monthrange(current_year, month_number)
    it_holidays = holidays.Italy(years=current_year)

    # --- Workbook Setup ---
    wb = Workbook()
    ws = wb.active
    ws.title = mese_str.capitalize()

    # --- Styles ---
    bold_font = Font(bold=True)
    light_red_fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    light_gold_fill = PatternFill(start_color="EEE8AA", end_color="EEE8AA", fill_type="solid")
    blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    light_purple_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")

    # --- Header ---
    ws.append(["Nome", nome, "Cognome", cognome])
    ws['B1'].fill = green_fill
    ws['D1'].fill = green_fill
    ws.append([])

    header_titles = ["Giorno", "Orario Mattina", "Orario Pomeriggio", "Permessi"]
    ws.append(header_titles)
    for cell in ws[ws.max_row]:
        cell.font = bold_font

    # --- Calendar Body ---
    for day_num in range(1, num_days + 1):
        current_date = date(current_year, month_number, day_num)
        date_str = current_date.strftime("%A, %d-%m-%Y").capitalize()
        
        # Default row data
        row_data = [date_str, "9:00 - 13:00", "14:00 - 18:00", ""]

        if current_date.weekday() >= 5 or current_date in it_holidays:
            row_data = [date_str, "", "", ""]
            ws.append(row_data)
            for cell in ws[ws.max_row]:
                cell.fill = light_red_fill
        
        elif day_num in giorni_ferie:
            row_data = [date_str, "FERIE", "", ""]
            ws.append(row_data)
            for cell in ws[ws.max_row]:
                cell.fill = light_gold_fill
            ws.cell(row=ws.max_row, column=2).font = bold_font

        elif day_num in giorni_malattia:
            row_data = [date_str, "MALATTIA", "", ""]
            ws.append(row_data)
            for cell in ws[ws.max_row]:
                cell.fill = blue_fill
            ws.cell(row=ws.max_row, column=2).font = bold_font

        elif day_num in permessi:
            permesso_info = permessi[day_num]
            row_data[3] = permesso_info["range"]
            ws.append(row_data)
            for cell in ws[ws.max_row]:
                cell.fill = light_purple_fill
        else:
            ws.append(row_data)
    
    # --- Totals ---
    total_permessi_hours = sum(p['hours'] for p in permessi.values())
    ws.append([]) # Spacer row
    
    ws.append(["Totale Ferie", len(giorni_ferie)])
    ws.cell(row=ws.max_row, column=1).font = bold_font
    
    ws.append(["Totale Malattia", len(giorni_malattia)])
    ws.cell(row=ws.max_row, column=1).font = bold_font

    ws.append(["Totale Ore Permesso", total_permessi_hours])
    ws.cell(row=ws.max_row, column=1).font = bold_font

    # --- Final Touches (Auto-fit columns) ---
    for col_idx, column in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # --- File Generation ---
    target = io.BytesIO()
    wb.save(target)
    target.seek(0)

    download_name = f"{nome}_{mese_str.capitalize()}{current_year}.xlsx"

    return send_file(
        target, 
        as_attachment=True, 
        download_name=download_name,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

def main():
    app.run(port=int(os.environ.get('PORT', 8080)))

if __name__ == "__main__":
    main()